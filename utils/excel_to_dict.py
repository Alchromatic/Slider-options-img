import json
from openpyxl import load_workbook

def parse_rgb_value(v):
    if not v:
        return None

    # "255,0,0" or "255, 0, 0"
    if isinstance(v, str) and "," in v:
        parts = v.split(",")
        if len(parts) == 3:
            return tuple(int(p.strip()) for p in parts)

    # "#FF0000"
    if isinstance(v, str) and v.startswith("#") and len(v) == 7:
        h = v[1:]
        return tuple(int(h[i:i+2], 16) for i in (0,2,4))

    return None


def extract_sheet(ws):
    colors = []

    # ---------------------
    # Check only row #2
    # ---------------------
    row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))

    rgb_col = None
    for idx, val in enumerate(row2):
        if isinstance(val, str) and val.strip().upper() == "RGB":
            rgb_col = idx
            break

    if rgb_col is None:
        return []

    # read from row 3 onwards
    for row in ws.iter_rows(min_row=3, values_only=True):
        rgb = parse_rgb_value(row[rgb_col])
        if rgb:
            colors.append(rgb)

    return colors


def excel_to_dict(path):
    wb = load_workbook(path, data_only=True)
    result = {}
    for sheet in wb.sheetnames:
        result[sheet] = extract_sheet(wb[sheet])
    return result


# ----------------------
# RUN
# ----------------------
file = "sunnysanwar.xlsx"
out = "colors.json"

data = excel_to_dict(file)

with open(out, "w") as f:
    json.dump(data, f, indent=4)

print("written", out)
